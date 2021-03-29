from openpyxl import *
from tkinter import *

wb = load_workbook('res/excelentry.xlsx')
sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Contact"
    sheet.cell(row=1, column=3).value = "Email"
    sheet.cell(row=1, column=4).value = "Gender"
    sheet.cell(row=1, column=5).value = "State"
    sheet.cell(row=1, column=6).value = "City"


def focus1(event):
    course_field.focus_set()


def focus2(event):
    # set focus on the sem_field box
    sem_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    form_no_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    address_field.focus_set()


root = Tk()


def getvalues():
    print("Form Submitted")
    print(f"Name: {namevalue.get()}, Contact No: {contactsvalue.get()}, Email Id: {emailvalue.get()}, Gender: {gendervalue.get()}, State: {statevalue.get()}, City: {cityvalue.get()}, Country; {nationality.get()}")


root.geometry("300x320")
root.title("Registration Form")


Label(root, text="Welcome Form", font="lucida 14 bold").grid(
    row=0, column=1, pady=15)

name = Label(root, text="Name").grid(row=1, column=0, padx=30, pady=5)
contact = Label(root, text="Contact").grid(row=2, column=0, padx=30, pady=5)
email = Label(root, text="Email").grid(row=3, column=0, padx=30, pady=5)
gender = Label(root, text="Gender").grid(row=4, column=0, padx=30, pady=5)
state = Label(root, text="State").grid(row=5, column=0, padx=30, pady=5)
city = Label(root, text="City").grid(row=6, column=0, padx=30, pady=5)

namevalue = StringVar()
contactsvalue = StringVar()
emailvalue = StringVar()
gendervalue = StringVar()
statevalue = StringVar()
cityvalue = StringVar()
nationality = IntVar()

user_entry = Entry(root, textvariable=namevalue).grid(row=1, column=1)
contact_entry = Entry(root, textvariable=contactsvalue).grid(row=2, column=1)
email_entry = Entry(root, textvariable=emailvalue).grid(row=3, column=1)
gender_entry = Entry(root, textvariable=gendervalue).grid(row=4, column=1)
state_entry = Entry(root, textvariable=statevalue).grid(row=5, column=1)
city_entry = Entry(root, textvariable=cityvalue).grid(row=6, column=1)

country = Checkbutton(
    text="Indian", variable=nationality).grid(row=7, column=1)

Button(root, text="Submit", command=getvalues).grid(row=8, column=1)

root.mainloop()
